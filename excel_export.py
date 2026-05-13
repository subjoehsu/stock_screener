"""
Excel export with colour-coded formatting.

Sheet layout:
  0 — 說明       (strategy summary)
  1 — 買點訊號   (buy signals, green theme)
  2 — 賣點訊號   (sell signals, red theme)
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────
_DARK_BLUE  = "1F4E79"
_DARK_GREEN = "1E6823"
_DARK_RED   = "9C0006"
_LIGHT_GREEN = "C6EFCE"
_LIGHT_RED   = "FFC7CE"
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_CENTER = Alignment(horizontal="center", vertical="center")

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _cell_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


# ── Sheet writer ──────────────────────────────────────────

def _write_data_sheet(
    ws,
    df: pd.DataFrame,
    sheet_title: str,
    header_color: str,
    check_cols: list[str],
) -> None:
    if df.empty:
        ws["A1"] = f"無 {sheet_title} 股票"
        ws["A1"].font = Font(italic=True, color="888888")
        return

    headers = list(df.columns)

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = _header_fill(header_color)
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, col_name in enumerate(headers, 1):
            value = row[col_name]
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER
            cell.border    = _BORDER

            if col_name in check_cols:
                if value == "✓":
                    cell.fill = _cell_fill(_LIGHT_GREEN)
                    cell.font = Font(color=_DARK_GREEN, bold=True)
                elif value == "✗":
                    cell.fill = _cell_fill(_LIGHT_RED)
                    cell.font = Font(color=_DARK_RED)

    # Auto column widths
    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max(
            len(str(h)),
            *(len(str(df.iloc[r][h])) for r in range(min(len(df), 50))),
        )
        ws.column_dimensions[col_letter].width = max(max_len * 1.4, 10)

    # Freeze top row
    ws.freeze_panes = "A2"


# ── Info sheet ────────────────────────────────────────────

def _write_info_sheet(ws, run_time: str) -> None:
    ws.title = "說明"
    ws.column_dimensions["A"].width = 50

    def row(r, text, bold=False, size=11, color="000000"):
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color)

    row(1,  "📈 SuperTREX 選股系統", bold=True, size=14, color=_DARK_BLUE)
    row(2,  f"分析時間：{run_time}")
    row(3,  "")
    row(4,  "【買點條件】（5項需全符合，或達到設定門檻）", bold=True, color=_DARK_GREEN)
    row(5,  "1. SuperTREX = Buy（趨勢翻多）")
    row(6,  "2. MACD 黃金交叉（近20根K棒內，fast=20 / slow=120 / signal=9）")
    row(7,  "3. RSI > 50（period=13）")
    row(8,  "4. 收盤價 > 20MA")
    row(9,  "5. 成交量 > 20日均量 × 1.5 倍")
    row(10, "")
    row(11, "【賣點條件】（4項需全符合，或達到設定門檻）", bold=True, color=_DARK_RED)
    row(12, "1. SuperTREX = Sell（趨勢翻空）")
    row(13, "2. MACD 死亡交叉（近20根K棒內）")
    row(14, "3. RSI < 45")
    row(15, "4. 收盤價 < 20MA（跌破）")
    row(16, "")
    row(17, "【SuperTREX 參數】", bold=True)
    row(18, "三組 SuperTrend 多數決：(14期, 4倍) + (35期, 4倍) + (70期, 10倍)")
    row(19, "")
    row(20, "【資料來源】")
    row(21, "台股：Yahoo Finance（股票代號.TW）")
    row(22, "美股：Yahoo Finance")
    row(23, "15分鐘K棒資料最長約59個曆日")


# ── Public API ────────────────────────────────────────────

BUY_CHECK_COLS  = ["SuperTREX=Buy",  "MACD黃金交叉", "RSI>50",  "收盤>MA20",  "成交量爆量×1.5"]
SELL_CHECK_COLS = ["SuperTREX=Sell", "MACD死亡交叉", "RSI<45", "跌破MA20"]


def build_excel(
    buy_df: pd.DataFrame,
    sell_df: pd.DataFrame,
    run_time: str | None = None,
) -> bytes:
    """
    Build an Excel workbook from screening results and return it as bytes.
    """
    if run_time is None:
        run_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # Remove the default empty sheet
    wb.remove(wb.active)

    # Info sheet (first tab)
    ws_info = wb.create_sheet("說明")
    _write_info_sheet(ws_info, run_time)

    # Buy sheet
    ws_buy = wb.create_sheet("買點訊號")
    ws_buy.title = "買點訊號"
    _write_data_sheet(
        ws_buy, buy_df,
        sheet_title="買點訊號",
        header_color=_DARK_GREEN,
        check_cols=BUY_CHECK_COLS,
    )

    # Sell sheet
    ws_sell = wb.create_sheet("賣點訊號")
    ws_sell.title = "賣點訊號"
    _write_data_sheet(
        ws_sell, sell_df,
        sheet_title="賣點訊號",
        header_color=_DARK_RED,
        check_cols=SELL_CHECK_COLS,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
